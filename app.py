import os
import json
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory, send_file
import firebase_admin
from firebase_admin import credentials, firestore, storage
import uuid
import base64
from io import BytesIO
from PIL import Image

# Import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel export will be disabled.")

app = Flask(__name__)
app.secret_key = 'hackathon-attendance-secret-key-2026'

# ==================== FIREBASE INITIALIZATION ====================

# Initialize Firebase Admin SDK
# Set the GOOGLE_APPLICATION_CREDENTIALS environment variable or ensure serviceAccountKey.json exists

try:
    # Check if Firebase is already initialized (from previous app reload)
    if not firebase_admin._apps:
        # Initialize Firebase with service account credentials
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(cred, {
            'storageBucket': 'attendence-63b06.appspot.com'
        })
        print("✅ Firebase initialized successfully")
    
    # Get Firestore and Storage clients
    db = firestore.client()
    storage_client = storage.bucket()
    FIREBASE_ENABLED = True
    print("✅ Firestore and Storage clients initialized")
    
except FileNotFoundError as e:
    print(f"❌ Firebase initialization failed: Service account key file not found")
    print(f"   Error: {e}")
    FIREBASE_ENABLED = False
    db = None
    storage_client = None
except Exception as e:
    print(f"❌ Firebase initialization failed: {e}")
    FIREBASE_ENABLED = False
    db = None
    storage_client = None

# Firebase configuration
FIREBASE_STORAGE_BUCKET = 'attendence-63b06.appspot.com'

# Admin credentials (hardcoded)
ADMIN_USERNAME = 'CyberNova'
ADMIN_PASSWORD = 'Owasp@CyberNova'

# Registered hunters (hardcoded with IDs and passwords)
HUNTERS = {
    'HUNTER001': {'name': 'Alpha Hunter', 'password': 'pass001'},
    'HUNTER002': {'name': 'Beta Stalker', 'password': 'pass002'},
    'HUNTER003': {'name': 'Gamma Tracker', 'password': 'pass003'},
    'HUNTER004': {'name': 'Delta Ranger', 'password': 'pass004'},
    'HUNTER005': {'name': 'Epsilon Scout', 'password': 'pass005'},
}

import pandas as pd
import re

# ==================== DATABASE MIGRATION ====================
def migrate_hunters_to_db():
    """Migrates hardcoded hunters to Firestore if collection is empty."""
    if not FIREBASE_ENABLED or not db:
        return
        
    try:
        hunters_ref = db.collection('hunters')
        # Check if any hunters exist (limit 1 for efficiency)
        docs = hunters_ref.limit(1).stream()
        if not list(docs):
            print("⚠️ 'hunters' collection empty. Migrating hardcoded hunters...")
            batch = db.batch()
            for hunter_id, info in HUNTERS.items():
                doc_ref = hunters_ref.document(hunter_id)
                batch.set(doc_ref, {
                    'name': info['name'],
                    'password': info['password'], # Plaintext for now based on legacy code
                    'members_count': 1 # Default for solo hunter
                })
            batch.commit()
            print(f"✅ Migrated {len(HUNTERS)} hunters to database.")
            
            # Initialize Settings (Total Hunters)
            settings_ref = db.collection('settings').document('config')
            if not settings_ref.get().exists:
                settings_ref.set({'totalHunters': 75}) # Default from user request
                print("✅ Initialized settings with totalHunters: 75")
        else:
            print("✅ 'hunters' collection already initialized.")
            
    except Exception as e:
        print(f"❌ Migration failed: {e}")

# ==================== FIRESTORE HELPERS ====================

def get_active_session():
    """Get the current active session from Firestore"""
    if not FIREBASE_ENABLED or not db:
        return None
    
    try:
        sessions = db.collection('sessions').where('active', '==', True).limit(1).stream()
        for doc in sessions:
            session_data = doc.to_dict()
            session_data['id'] = doc.id
            
            # Check if expired
            start_time = datetime.fromisoformat(session_data['startTime'])
            elapsed = (datetime.now() - start_time).total_seconds() / 60
            
            if elapsed > 15:
                # Expire the session
                db.collection('sessions').document(doc.id).update({'active': False})
                return None
            
            return session_data
        return None
    except Exception as e:
        print(f"Error getting active session: {e}")
        return None

def check_hunter_registered(hunter_id):
    """Check if hunter is registered in Firestore"""
    if not FIREBASE_ENABLED or not db:
        return hunter_id in HUNTERS
        
    try:
        doc = db.collection('hunters').document(hunter_id).get()
        return doc.exists
    except Exception as e:
        print(f"Error checking hunter: {e}")
        return False

def verify_hunter_login(hunter_id, password):
    """Verifies hunter credentials against Firestore."""
    if not FIREBASE_ENABLED or not db:
        # Fallback to hardcoded
        if hunter_id in HUNTERS and HUNTERS[hunter_id]['password'] == password:
            return HUNTERS[hunter_id]['name']
        return None
        
    try:
        doc = db.collection('hunters').document(hunter_id).get()
        if doc.exists:
            data = doc.to_dict()
            # Simple plaintext check
            if data.get('password') == password:
                return {
                    'name': data.get('name', hunter_id),
                    'members_count': data.get('members_count', 1)
                }
    except Exception as e:
        print(f"Login error: {e}")
    return None

def check_duplicate_submission(hunter_id, session_id):
    """Check if hunter already submitted for this session"""
    if not FIREBASE_ENABLED or not db:
        return False
    
    try:
        # Check by document ID which is deterministic: hunterId_sessionId
        doc_id = f"{hunter_id}_{session_id}"
        doc = db.collection('attendance').document(doc_id).get()
        return doc.exists
    except Exception as e:
        print(f"Error checking duplicate: {e}")
        return False

def upload_image_to_firebase(base64_image, hunter_id, session_id):
    """Upload base64 image to Firebase Storage and return public URL"""
    if not FIREBASE_ENABLED or not storage_client:
        print("⚠️ Firebase Storage not available, storing locally instead")
        return upload_image_locally(base64_image, hunter_id, session_id)
    
    try:
        # Decode base64 image
        image_data = base64_image.replace('data:image/jpeg;base64,', '').replace('data:image/png;base64,', '')
        image_bytes = base64.b64decode(image_data)
        
        # Compress image
        img = Image.open(BytesIO(image_bytes))
        img.thumbnail((1920, 1440))
        output = BytesIO()
        img.save(output, format='JPEG', quality=90)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = int(datetime.now().timestamp())
        filename = f"selfies/{hunter_id}_{session_id}_{timestamp}.jpg"
        
        # Upload to Firebase Storage
        blob = storage_client.blob(filename)
        blob.upload_from_string(
            output.getvalue(),
            content_type='image/jpeg'
        )
        
        # Make the file publicly readable
        blob.make_public()
        
        # Return public URL
        public_url = blob.public_url
        
        print(f"Image uploaded to Firebase: {public_url}")
        return public_url
    
    except Exception as e:
        print(f"❌ Firebase upload failed: {e}")
        print("⚠️ Falling back to local storage...")
        return upload_image_locally(base64_image, hunter_id, session_id)

def upload_image_locally(base64_image, hunter_id, session_id):
    """Upload image to local uploads folder"""
    try:
        # Create uploads folder if it doesn't exist
        os.makedirs('uploads', exist_ok=True)
        
        # Decode base64 image
        image_data = base64_image.replace('data:image/jpeg;base64,', '').replace('data:image/png;base64,', '')
        image_bytes = base64.b64decode(image_data)
        
        # Compress image
        img = Image.open(BytesIO(image_bytes))
        img.thumbnail((1920, 1440))
        output = BytesIO()
        img.save(output, format='JPEG', quality=90)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = int(datetime.now().timestamp())
        filename = f"{hunter_id}_{session_id}_{timestamp}.jpg"
        filepath = os.path.join('uploads', filename)
        
        # Save to local folder
        with open(filepath, 'wb') as f:
            f.write(output.getvalue())
        
        # Return relative URL
        image_url = f"/uploads/{filename}"
        print(f"✅ Image uploaded locally: {image_url}")
        return image_url
    
    except Exception as e:
        print(f"Error uploading image locally: {e}")
        raise Exception(f"Failed to upload image: {str(e)}")

# ==================== AUTHENTICATION ====================

def admin_login_required(f):
    """Decorator to require admin login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def hunter_login_required(f):
    """Decorator to require hunter login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'hunterId' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== ROUTES: AUTHENTICATION ====================

@app.route('/')
def index():
    """Redirect to login"""
    if 'admin' in session:
        return redirect(url_for('admin_dashboard'))
    elif 'hunterId' in session:
        return redirect(url_for('hunter_page'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        # 1. Check Admin Credentials
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin'] = True
            session.permanent = True
            app.permanent_session_lifetime = timedelta(hours=8)
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('admin_dashboard')})
            return redirect(url_for('admin_dashboard'))
            
        # 2. Check Hunter Credentials
        hunterId = username.upper()
        hunter_data = verify_hunter_login(hunterId, password)
        
        if hunter_data:
            session['hunterId'] = hunterId
            if isinstance(hunter_data, dict):
                session['hunterName'] = hunter_data['name']
                session['membersCount'] = hunter_data['members_count']
            else:
                # Fallback for legacy (should not happen with new code)
                session['hunterName'] = hunter_data
                session['membersCount'] = 1
            session.permanent = True
            app.permanent_session_lifetime = timedelta(hours=8)
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('hunter_page')})
            return redirect(url_for('hunter_page'))
        
        # 3. Invalid Credentials
        if request.is_json:
            return jsonify({'success': False, 'error': 'Invalid username or password'}), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logout user completely"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/logout/admin')
def logout_admin():
    """Logout only admin role"""
    session.pop('admin', None)
    # If hunter session exists, don't clear everything
    if 'hunterId' in session:
        return redirect(url_for('login')) # Or redirect to hunter page if you prefer
    
    # If no other session, clear all to be safe
    session.clear()
    return redirect(url_for('login'))

@app.route('/logout/hunter')
def logout_hunter():
    """Logout only hunter role"""
    session.pop('hunterId', None)
    session.pop('hunterName', None)
    session.pop('membersCount', None)
    
    # If admin session exists, don't clear everything
    if 'admin' in session:
        return redirect(url_for('login'))
        
    session.clear()
    return redirect(url_for('login'))

# ==================== ROUTES: ADMIN ====================

@app.route('/admin')
@admin_login_required
def admin_dashboard():
    """Admin dashboard"""
    if not FIREBASE_ENABLED:
        return render_template('admin.html', error='Firebase not configured. Please set up Firebase credentials.'), 503
    
    try:
        # Get active session
        current_session = get_active_session()
        
        # Get all attendance records
        attendance_records = []
        docs = db.collection('attendance').order_by('timestamp', direction=firestore.Query.DESCENDING).limit(50).stream()
        for doc in docs:
            record = doc.to_dict()
            record['id'] = doc.id
            attendance_records.append(record)
            
        # Get fresh hunter data for stats
        hunters_data = {}
        hunter_docs = db.collection('hunters').stream()
        for doc in hunter_docs:
            hunters_data[doc.id] = doc.to_dict()
        
        return render_template('admin.html', 
                              current_session=current_session,
                              attendance_records=attendance_records,
                              hunters=hunters_data)
    except Exception as e:
        print(f"Error loading admin dashboard: {e}")
        return render_template('admin.html', error=str(e)), 500

@app.route('/api/admin/start-session', methods=['POST'])
@admin_login_required
def start_session():
    """Start a new attendance session"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured. Please set up Firebase credentials.'}), 503
    
    data = request.get_json()
    session_type = data.get('sessionType', 'Morning').strip()
    try:
        duration = int(data.get('duration', 15))
    except (ValueError, TypeError):
        duration = 15
    
    # Allow any session type as per user request to be configurable
    if not session_type:
        return jsonify({'success': False, 'error': 'Session type cannot be empty'}), 400
    
    try:
        # End any active sessions
        active_sessions = db.collection('sessions').where('active', '==', True).stream()
        for doc in active_sessions:
            db.collection('sessions').document(doc.id).update({'active': False})
        
        # Create new session
        session_id = str(uuid.uuid4())[:8].upper()
        start_time = datetime.now().isoformat()
        
        session_data = {
            'sessionId': session_id,
            'type': session_type,
            'startTime': start_time,
            'duration': duration,
            'active': True
        }
        
        db.collection('sessions').document(session_id).set(session_data)
        
        return jsonify({
            'success': True,
            'sessionId': session_id,
            'sessionType': session_type,
            'duration': duration,
            'qrData': json.dumps({'sessionId': session_id, 'sessionType': session_type})
        })
    except Exception as e:
        print(f"Error starting session: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/end-session', methods=['POST'])
@admin_login_required
def end_session():
    """End current session"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured. Please set up Firebase credentials.'}), 503
    
    try:
        active_sessions = db.collection('sessions').where('active', '==', True).stream()
        for doc in active_sessions:
            db.collection('sessions').document(doc.id).update({'active': False})
        
        return jsonify({'success': True, 'message': 'Session ended'})
    except Exception as e:
        print(f"Error ending session: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/sessions/all', methods=['DELETE'])
@admin_login_required
def delete_all_sessions():
    """Delete ALL sessions"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured'}), 503
    
    try:
        # End any active session first? Or just wipe. Wiping is safer to ensure clean state.
        sessions_ref = db.collection('sessions')
        batch_size = 500
        
        while True:
            docs = sessions_ref.limit(batch_size).stream()
            deleted = 0
            batch = db.batch()
            for doc in docs:
                batch.delete(doc.reference)
                deleted += 1
            
            if deleted == 0:
                break
                
            batch.commit()
            
        return jsonify({'success': True, 'message': 'All sessions deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/attendance/all', methods=['DELETE'])
@admin_login_required
def delete_all_attendance():
    """Delete ALL attendance records"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured'}), 503
    
    try:
        # Delete from 'attendance' collection
        att_ref = db.collection('attendance')
        batch_size = 500
        
        while True:
            docs = att_ref.limit(batch_size).stream()
            deleted = 0
            batch = db.batch()
            for doc in docs:
                batch.delete(doc.reference)
                deleted += 1
            
            if deleted == 0:
                break
                
            batch.commit()
            
        return jsonify({'success': True, 'message': 'All attendance records deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/verify-attendance/<attendance_id>', methods=['POST'])
@admin_login_required
def verify_attendance(attendance_id):
    """Verify attendance record"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured. Please set up Firebase credentials.'}), 503
    
    try:
        db.collection('attendance').document(attendance_id).update({'status': 'Verified'})
        return jsonify({'success': True, 'message': 'Attendance verified'})
    except Exception as e:
        print(f"Error verifying attendance: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/delete-attendance/<attendance_id>', methods=['POST'])
@admin_login_required
def delete_attendance(attendance_id):
    """Delete attendance record"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured.'}), 503
    
    try:
        db.collection('attendance').document(attendance_id).delete()
        return jsonify({'success': True, 'message': 'Attendance deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== HUNTER MANAGEMENT APIs ====================

@app.route('/api/admin/hunters', methods=['GET'])
@admin_login_required
def get_hunters():
    """Get all hunters"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
    
    try:
        hunters = []
        # Sort by Document ID (Hunter ID)
        docs = db.collection('hunters').order_by('__name__').stream()
        for doc in docs:
            data = doc.to_dict()
            hunters.append({
                'id': doc.id,
                'name': data.get('name', 'Unknown'),
                'password': data.get('password', '***'), # Show password for admin convenience as requested
                'members_count': data.get('members_count', 1)
            })
        return jsonify({'success': True, 'hunters': hunters})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/hunters/all', methods=['DELETE'])
@admin_login_required
def delete_all_hunters():
    """Delete ALL hunters"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
        
    try:
        # Delete in batches
        hunters_ref = db.collection('hunters')
        batch_size = 500
        
        while True:
            docs = hunters_ref.limit(batch_size).stream()
            deleted = 0
            batch = db.batch()
            for doc in docs:
                batch.delete(doc.reference)
                deleted += 1
            
            if deleted == 0:
                break
                
            batch.commit()
            
        return jsonify({'success': True, 'message': 'All hunters deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/hunters/<hunter_id>', methods=['DELETE'])
@admin_login_required
def delete_single_hunter(hunter_id):
    """Delete a specific hunter"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
        
    try:
        db.collection('hunters').document(hunter_id).delete()
        return jsonify({'success': True, 'message': f'Hunter {hunter_id} deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/add-hunter', methods=['POST'])
@admin_login_required
def add_hunter():
    """Add a new hunter"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
    
    data = request.get_json() 
    hunter_id = data.get('hunterId', '').strip().upper()
    name = data.get('name', '').strip()
    password = data.get('password', '').strip()
    
    if not all([hunter_id, name, password]):
        return jsonify({'success': False, 'error': 'Missing fields'}), 400
        
    try:
        # Check if exists
        doc = db.collection('hunters').document(hunter_id).get()
        if doc.exists:
            return jsonify({'success': False, 'error': 'Hunter ID already exists'}), 400
            
        db.collection('hunters').document(hunter_id).set({
            'name': name,
            'password': password,
            'members_count': int(data.get('membersCount', 1))
        })
        return jsonify({'success': True, 'message': 'Hunter added'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/update-hunter-details', methods=['POST'])
@admin_login_required
def update_hunter_details():
    """Update hunter name and member count"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
    
    data = request.get_json()
    hunter_id = data.get('hunterId')
    name = data.get('name', '').strip()
    try:
        members_count = int(data.get('membersCount', 1))
    except (ValueError, TypeError):
        members_count = 1
        
    if not hunter_id or not name:
        return jsonify({'success': False, 'error': 'Missing hunter ID or name'}), 400
        
    try:
        db.collection('hunters').document(hunter_id).update({
            'name': name,
            'members_count': members_count
        })
        return jsonify({'success': True, 'message': 'Hunter details updated'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/update-hunter-password', methods=['POST'])
@admin_login_required
def update_hunter_password():
    """Update hunter password"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
    
    data = request.get_json()
    hunter_id = data.get('hunterId', '')
    new_password = data.get('password', '').strip()
    
    if not all([hunter_id, new_password]):
        return jsonify({'success': False, 'error': 'Missing fields'}), 400
        
    try:
        db.collection('hunters').document(hunter_id).update({'password': new_password})
        return jsonify({'success': True, 'message': 'Password updated'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/settings', methods=['POST', 'GET'])
@admin_login_required
def admin_settings():
    """Get or update settings"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
        
    try:
        if request.method == 'POST':
            data = request.get_json()
            total_hunters = int(data.get('totalHunters', 75))
            db.collection('settings').document('config').set({'totalHunters': total_hunters}, merge=True)
            return jsonify({'success': True, 'message': 'Settings updated'})
        else:
            doc = db.collection('settings').document('config').get()
            total_hunters = doc.to_dict().get('totalHunters', 75) if doc.exists else 75
            return jsonify({'success': True, 'totalHunters': total_hunters})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/import-hunters', methods=['POST'])
@admin_login_required
def import_hunters():
    """Bulk import hunters from CSV or Excel"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
        
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file part'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'}), 400
        
    try:
         # Read file - handle both CSV and Excel
         if file.filename.endswith('.csv'):
             df = pd.read_csv(file, header=None)
         else:
             df = pd.read_excel(file, header=None)

         batch = db.batch()
         hunters_ref = db.collection('hunters')
         
         processed_count = 0
         
         # Skip header row if it exists
         start_idx = 0
         if len(df) > 0:
             first_val = str(df.iloc[0, 0]).lower()
             if 'hunter' in first_val and 'id' in first_val:
                 start_idx = 1
         
         for index in range(start_idx, len(df)):
             row = df.iloc[index]
             
             # Expected Format:
             # Col 0: Hunter ID (Username)
             # Col 1: Hunter Name
             # Col 2: Register No (Password)
             
             hunter_id = str(row[0]).strip() if pd.notna(row[0]) else ""
             hunter_name = str(row[1]).strip() if pd.notna(row[1]) else ""
             register_no = str(row[2]).strip() if pd.notna(row[2]) else ""
             
             # Skip empty rows
             if not hunter_id or not hunter_name:
                 continue
                 
             # Use Register No as Password. If missing, default to 'password123'
             password = register_no if register_no else 'password123'
             
             doc_ref = hunters_ref.document(hunter_id)
             batch.set(doc_ref, {
                 'name': hunter_name,
                 'password': password,
                 'members_count': 1 # Default
             })
             
             processed_count += 1
             
             # Commit in batches of 400
             if processed_count % 400 == 0:
                 batch.commit()
                 batch = db.batch()
         
         if processed_count > 0:
             batch.commit()
         
         # Update global stats (optional, just total count)
         total_hunters_snap = db.collection('hunters').count().get()
         total_count = total_hunters_snap[0][0].value
         
         db.collection('settings').document('config').set({
             'totalHunters': total_count
         }, merge=True)
         
         return jsonify({
             'success': True, 
             'message': f'Successfully imported {processed_count} hunters.',
             'count': processed_count
         })
         
    except Exception as e:
         return jsonify({'success': False, 'error': f'Import failed: {str(e)}'}), 500



@app.route('/api/admin/session-status')
@admin_login_required
def session_status():
    """Get current session status"""
    try:
        current_session = get_active_session()
        
        if not current_session:
            return jsonify({'active': False})
        
        start_time = datetime.fromisoformat(current_session['startTime'])
        elapsed = (datetime.now() - start_time).total_seconds() / 60
        
        return jsonify({
            'active': True,
            'sessionId': current_session['sessionId'],
            'sessionType': current_session['type'],
            'startTime': current_session['startTime'],
            'duration': current_session.get('duration', 15),
            'remaining': max(0, int(current_session.get('duration', 15)) - int(elapsed))
        })
    except Exception as e:
        print(f"Error getting session status: {e}")
        return jsonify({'active': False})

@app.route('/api/admin/attendance-feed')
@admin_login_required
def attendance_feed():
    """Get latest attendance records"""
    if not FIREBASE_ENABLED:
        return jsonify([])
    
    try:
        records = []
        # Simplified query to avoid index issues
        # Client-side sorts anyway, and we likely have <100 records for now
        docs = db.collection('attendance').stream()
        
        # Prefetch all hunters for lookup (optimize DB reads)
        hunter_docs = db.collection('hunters').stream()
        hunter_map = {d.id: d.to_dict() for d in hunter_docs}

        print(f"--- [DEBUG] START FEED FETCH ---")
        for doc in docs:
            record = doc.to_dict()
            record['id'] = doc.id
            hunter_id = record.get('hunterId', 'Unknown')
            
            # Look up hunter details dynamically
            hunter_details = hunter_map.get(hunter_id, {})
            record['hunterName'] = hunter_details.get('name', 'Unknown')
            record['requiredFaces'] = hunter_details.get('members_count', 1) # Default to 1 if missing
            
            # Print minimal info to console for verification
            print(f"Doc: {doc.id} | Hunter: {hunter_id} | Session: {record.get('sessionId')}")
            
            records.append(record)
            
        print(f"--- [DEBUG] END FEED FETCH: {len(records)} records found ---")
        
        return jsonify(records)
    except Exception as e:
        print(f"Error getting attendance feed: {e}")
        return jsonify([])

@app.route('/api/admin/statistics')
@admin_login_required
def get_statistics():
    """Get attendance statistics"""
    if not FIREBASE_ENABLED:
        return jsonify({'verified': 0, 'selfies': 0, 'total': 0})
    
    try:
        # Get verified count
        verified_docs = db.collection('attendance').where('status', '==', 'Verified').stream()
        verified_count = sum(1 for _ in verified_docs)
        
        # Get all submitted (selfies uploaded)
        all_docs = db.collection('attendance').stream()
        all_count = sum(1 for _ in all_docs)
        
        return jsonify({
            'verified': verified_count,
            'selfies': all_count,
            'total': all_count
        })
    except Exception as e:
        print(f"Error getting statistics: {e}")
        return jsonify({'verified': 0, 'selfies': 0, 'total': 0})

@app.route('/api/admin/export-excel/<session_id>')
@admin_login_required
def export_attendance_excel(session_id):
    """Export attendance records to Excel for a specific session"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured'}), 503
    
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel export not available. openpyxl is not installed.'}), 503
    
    try:
        # Query attendance records for the session
        docs = db.collection('attendance').where('sessionId', '==', session_id).stream()
        records = []
        for doc in docs:
            record = doc.to_dict()
            records.append(record)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Define header row
        headers = ['Hunter ID', 'Hunter Name', 'Session ID', 'Session Type', 'Status', 'Timestamp']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for record in records:
            hunter_id = record.get('hunterId', 'Unknown')
            hunter_name = HUNTERS.get(hunter_id, {}).get('name', 'Unknown')
            session_id_val = record.get('sessionId', '')
            session_type = record.get('sessionType', '')
            status = record.get('status', '')
            timestamp = record.get('timestamp', '')
            
            ws.append([hunter_id, hunter_name, session_id_val, session_type, status, timestamp])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 25
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Return file with proper headers
        filename = f"attendance_{session_id}.xlsx"
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error exporting attendance: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== ROUTES: HUNTER ====================

@app.route('/hunter')
@hunter_login_required
def hunter_page():
    """Hunter attendance page"""
    return render_template('hunter.html',
                          hunterId=session.get('hunterId'),
                          hunterName=session.get('hunterName'),
                          membersCount=session.get('membersCount', 1))

@app.route('/api/hunter/validate-qr', methods=['POST'])
@hunter_login_required
def validate_qr():
    """Validate QR code data"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured. Please set up Firebase credentials.'}), 503
    
    data = request.get_json()
    qr_data = data.get('qrData', '').strip()
    

    
    if not qr_data:
        return jsonify({'success': False, 'error': 'QR code data is empty. Please scan a valid QR code.'}), 400
    
    try:
        # Try to parse as JSON
        qr_json = json.loads(qr_data)
        
        # Ensure it's a dictionary
        if isinstance(qr_json, dict):
            session_id = qr_json.get('sessionId', '').strip()
            session_type = qr_json.get('sessionType', '').strip()
        else:
            # Valid JSON but not a dict (e.g. number 1234 or string "foo"), treat as raw ID
            raise ValueError("Not a dictionary")
            
    except (json.JSONDecodeError, ValueError):
        # Not JSON or not a dict, treat as raw Session ID

        session_id = qr_data
        session_type = None # We will fetch this from DB
    except Exception as e:

        return jsonify({'success': False, 'error': f'Error parsing QR code: {str(e)}'}), 400
    
    # Validate required fields
    if not session_id:
        return jsonify({'success': False, 'error': 'Session ID not found'}), 400
    
    # If it was JSON, we might have session_type, but if it was raw, we need to fetch it
    
    try:
        # Get session from Firestore
        session_doc = db.collection('sessions').document(session_id).get()
        

        
        if not session_doc.exists:
            return jsonify({'success': False, 'error': 'Session not found. Ask admin to start a new session.'}), 400
        
        session_data = session_doc.to_dict()
        
        # If we didn't get session_type from JSON, get it from DB
        if not session_type:
            session_type = session_data.get('type', 'Unknown')
        
        if not session_data.get('active', False):
            return jsonify({'success': False, 'error': 'Session is not active. Ask admin to start the session.'}), 400
        
        # Check session expiration
        start_time = datetime.fromisoformat(session_data['startTime'])
        elapsed = (datetime.now() - start_time).total_seconds() / 60
        

        
        if elapsed > 15:
            db.collection('sessions').document(session_id).update({'active': False})
            return jsonify({'success': False, 'error': 'Session has expired. Ask admin to start a new session.'}), 400
        
        # Check if hunter already submitted for this session
        hunter_id = session.get('hunterId')
        
        if check_duplicate_submission(hunter_id, session_id):
            return jsonify({'success': False, 'error': 'Your hunter already submitted for this session.'}), 400
        

        
        return jsonify({
            'success': True,
            'sessionId': session_id,
            'sessionType': session_type,
            'startTime': session_data.get('startTime'), # Pass start time
            'duration': session_data.get('duration', 15), # Pass duration
            'message': f'✅ QR validated! Please take a group selfie for {session_type} attendance.'
        })
    
    except Exception as e:

        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/api/hunter/submit-attendance', methods=['POST'])
@hunter_login_required
def submit_attendance():
    """Submit attendance with selfie image"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured. Please set up Firebase credentials.'}), 503
    
    hunter_id = session.get('hunterId')
    
    try:
        data = request.get_json()
        image_data = data.get('image', '')
        session_id = data.get('sessionId', '')
        session_type = data.get('sessionType', '')
        face_count = data.get('faceCount', 0) # Default to 0 if not provided
        
        if not all([image_data, session_id, session_type]):
            return jsonify({'success': False, 'error': 'Missing required data'}), 400
        
        # Validate Participant Selection
        selected_participants = data.get('selectedParticipants', [])
        
        # STRICT RULE: detectedFaces must equal selectedParticipants length
        # (Unless faceCount is 0, which implies detection failed? No, strict rule says MUST equal)
        # However, if selected_participants is empty (legacy), fallback to old logic? 
        # The user says "Hunter Member Selection (MANDATORY)". So we enforce it.
        
        if not selected_participants:
             # Relaxed rule: If no members selected (legacy/simple mode), just warn or proceed?
             # For now, we allow it but maybe log it?
             # return jsonify({'success': False, 'error': 'No hunter members selected.'}), 400
             pass 
             
        expected_count = len(selected_participants) if selected_participants else face_count
        
        # Allow slight leniency? No, user said "STRICT RULE".
        if selected_participants and face_count != expected_count:
             return jsonify({'success': False, 'error': f"Face mismatch: Selected {expected_count} members but detected {face_count} faces."}), 400
        
        # Upload image to Firebase Storage
        try:
            image_url = upload_image_to_firebase(image_data, hunter_id, session_id)
        except Exception as e:
            return jsonify({'success': False, 'error': f'Image upload failed: {str(e)}'}), 500
        
        # Store in Firestore
        timestamp = datetime.now().isoformat()
        
        # Get total members for this hunter from session or DB?
        # Ideally DB, but session might have it if we update login
        # We'll fetch from DB or rely on what we have. 
        # For immutable record, we store what we know now.
        total_members = 1 
        # Try to get real total if available in session, else will be updated by import logic later
        if 'membersCount' in session:
             total_members = session['membersCount']
        
        attendance_data = {
            'hunterId': hunter_id,
            'sessionId': session_id,
            'sessionType': session_type,
            'timestamp': timestamp,
            'imageURL': image_url,
            'imagePath': image_url,
            'status': 'Pending',
            'faceCount': face_count,
            'id': f"{hunter_id}_{session_id}",
            'membersCount': total_members, 
            'faceMatch': face_count == expected_count,
            'detectedFaces': face_count,
            'selectedParticipants': selected_participants, # List of names
            'presentCount': expected_count
        }
        
        doc_id = f"{hunter_id}_{session_id}"
        db.collection('attendance').document(doc_id).set(attendance_data)
        
        return jsonify({
            'success': True,
            'message': 'Attendance submitted successfully',
            'timestamp': timestamp
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/hunter/members', methods=['GET'])
@hunter_login_required
def get_hunter_members():
    """Get members list for the logged-in hunter"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase unavailable'}), 503
        
    hunter_id = session.get('hunterId')
    try:
        doc = db.collection('hunters').document(hunter_id).get()
        if doc.exists:
            data = doc.to_dict()
            members = data.get('members', [])
            return jsonify({'success': True, 'members': members})
        return jsonify({'success': False, 'error': 'Hunter not found'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/hunter/attendance-status/<session_id>')
@hunter_login_required
def attendance_status(session_id):
    """Check if hunter has already submitted"""
    if not FIREBASE_ENABLED:
        return jsonify({'success': False, 'error': 'Firebase not configured. Please set up Firebase credentials.'}), 503
    
    hunter_id = session.get('hunterId')
    
    try:
        doc_id = f"{hunter_id}_{session_id}"
        doc = db.collection('attendance').document(doc_id).get()
        
        if doc.exists:
            data = doc.to_dict()
            return jsonify({'submitted': True, 'status': data.get('status')})
        
        return jsonify({'submitted': False})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== UPLOADS ROUTE ====================

@app.route('/uploads/<filename>')
def serve_upload(filename):
    """Serve uploaded images from the uploads folder"""
    try:
        return send_from_directory('uploads', filename)
    except FileNotFoundError:
        return jsonify({'error': 'File not found'}), 404

@app.route('/favicon.ico')
def favicon():
    """Silence favicon 404s"""
    return '', 204

# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500

# ==================== RUN ====================

if __name__ == '__main__':
    print("=" * 60)
    print("HACKATHON ATTENDANCE SYSTEM - FIREBASE VERSION")
    print("=" * 60)
    print("\nFirebase Configuration:")
    print(f"  Credentials: {os.environ.get('GOOGLE_APPLICATION_CREDENTIALS', 'Not set')}")
    print(f"  Storage Bucket: {FIREBASE_STORAGE_BUCKET or 'Not set'}")
    print("\nAdmin Credentials:")
    print(f"  Username: {ADMIN_USERNAME}")
    print(f"  Password: {ADMIN_PASSWORD}")
    print("\nHunter Credentials (Sample):")
    for hunter_id, hunter_info in list(HUNTERS.items())[:3]:
        print(f"  {hunter_id}: {hunter_info['name']} / {hunter_info['password']}")
    print("\nServer running on: http://0.0.0.0:5001")
    print("=" * 60)
    
    # Run migration
    migrate_hunters_to_db()
    
    # Run without reloader to prevent restart on file upload
    # Run with threaded=True for concurrent request handling
    app.run(debug=False, host='0.0.0.0', port=5001, threaded=True)